#!/usr/bin/env python3
"""
اسکریپت تکمیل اطلاعات فیلم‌های 9_Movies.xlsx
از 1% تا 100% با قابلیت ادامه از نقطه قطع شده

نحوه استفاده:
1. این فایل را در کنار 9_Movies.xlsx قرار دهید
2. اجرا کنید: python enrich_9_movies.py
3. برای ادامه از نقطه قطع شده، دوباره اجرا کنید
"""

import json
import requests
import time
import openpyxl
import re
from pathlib import Path

API_KEY = "fe53f97e"
BASE_URL = "https://www.omdbapi.com/"
PROGRESS_FILE = Path('progress.json')

def fetch_movie_info(movie_name):
    """دریافت اطلاعات فیلم از OMDb API"""
    try:
        # حذف سال و اطلاعات اضافی
        clean_name = re.sub(r'\s*\(\d{4}\)\s*$', '', movie_name).strip()
        clean_name = re.sub(r'\s*-\s*.*$', '', clean_name).strip()
        
        url = f"{BASE_URL}?t={requests.utils.quote(clean_name)}&apikey={API_KEY}"
        response = requests.get(url, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            if data.get('Response') == 'True':
                return {
                    'Title': data.get('Title', ''),
                    'Original Title': data.get('Title', ''),
                    'Year': data.get('Year', ''),
                    'Genre': data.get('Genre', ''),
                    'Director': data.get('Director', ''),
                    'Cast': data.get('Actors', ''),
                    'Rating': data.get('imdbRating', ''),
                    'Runtime': data.get('Runtime', '').replace(' min', ''),
                    'Country': data.get('Country', ''),
                    'Synopsis': data.get('Plot', ''),
                    'Poster URL': data.get('Poster', '')
                }
    except Exception as e:
        print(f"  خطا: {e}")
    return None

def enrich_movies():
    """تکمیل اطلاعات فیلم‌ها از 1% تا 100%"""
    
    print("="*60)
    print("شروع تکمیل اطلاعات فیلم‌ها")
    print("="*60)
    
    # باز کردن فایل
    input_file = '9_Movies.xlsx'
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    # پیدا کردن ستون‌ها
    headers = [cell.value for cell in ws[1]]
    
    col_map = {}
    for col_name in ['Title', 'Original Title', 'Director', 'Cast', 'Year', 'Genre', 'Rating', 'Runtime', 'Country', 'Synopsis', 'Poster URL']:
        if col_name in headers:
            col_map[col_name] = headers.index(col_name)
    
    if 'Title' not in col_map:
        print("خطا: ستون Title وجود ندارد")
        return
    
    total_rows = ws.max_row - 1
    print(f"تعداد کل فیلم‌ها: {total_rows}")
    
    # بارگذاری پیشرفت
    start_row = 2
    enriched_count = 0
    
    if PROGRESS_FILE.exists():
        try:
            with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
                progress = json.load(f)
                start_row = progress.get('last_row', 2)
                enriched_count = progress.get('results_count', 0)
                print(f"ادامه از ردیف {start_row} ({enriched_count} تکمیل شده)")
        except:
            pass
    
    # پردازش
    processed = 0
    enriched = enriched_count
    
    for row_idx in range(start_row, ws.max_row + 1):
        # بررسی پردازش قبلی
        title = ws.cell(row=row_idx, column=col_map['Title'] + 1).value
        if title and str(title).strip() != '':
            continue
        
        # استخراج نام فیلم
        path = ws.cell(row=row_idx, column=2).value
        if not path:
            continue
        
        match = re.search(r'\\([^\\]+)$', str(path))
        if not match:
            continue
        
        folder_name = match.group(1)
        clean_name = re.sub(r'^\d+', '', folder_name).strip()
        
        # نمایش پیشرفت
        progress_pct = (row_idx / (ws.max_row - 1)) * 100
        print(f"[{progress_pct:.1f}%] {clean_name}", end=" ")
        
        # دریافت اطلاعات
        info = fetch_movie_info(clean_name)
        
        if info:
            for col_name, col_idx in col_map.items():
                if col_name in info:
                    ws.cell(row=row_idx, column=col_idx + 1, value=info[col_name])
            
            print("✓")
            enriched += 1
        else:
            print("✗")
        
        processed += 1
        
        # ذخیره پیشرفت هر 10 فیلم
        if processed % 10 == 0:
            wb.save(input_file)
            with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
                json.dump({
                    'last_row': row_idx + 1,
                    'results_count': enriched,
                    'processed': processed,
                    'total': total_rows
                }, f, ensure_ascii=False)
            print(f"  → {enriched}/{total_rows} ({(enriched/total_rows*100):.1f}%)")
        
        # احترام به rate limit
        time.sleep(0.3)
    
    # ذخیره نهایی
    wb.save(input_file)
    
    # گزارش
    print(f"\n{'='*60}")
    print(f"✅ تکمیل شد!")
    print(f"{'='*60}")
    print(f"کل فیلم‌ها: {total_rows}")
    print(f"پردازش شده: {processed}")
    print(f"تکمیل شده: {enriched}")
    print(f"درصد تکمیل: {(enriched/total_rows*100):.1f}%")
    print(f"{'='*60}")

if __name__ == '__main__':
    enrich_movies()
