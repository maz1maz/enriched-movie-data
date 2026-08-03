#!/usr/bin/env python3
"""
اسکریپت تکمیل اطلاعات فیلم‌های فایل 9_Movies.xlsx
با قابلیت ادامه از نقطه قطع شده

نحوه استفاده:
1. این فایل را در کنار فایل‌های 9_Movies_part1.xlsx تا 9_Movies_part10.xlsx قرار دهید
2. اجرا کنید: python enrich_9_movies_complete.py
3. برای ادامه از نقطه قطع شده، دوباره اجرا کنید
4. پس از تکمیل همه قسمت‌ها، فایل اصلی 9_Movies.xlsx به‌روزرسانی می‌شود

API Key: fe53f97e (OMDb)
"""

import json
import requests
import time
import openpyxl
import re
import os
import sys
from pathlib import Path
from urllib.parse import quote

# ============== تنظیمات ==============
API_KEY = "fe53f97e"
BASE_URL = "https://www.omdbapi.com/"
RATE_LIMIT_DELAY = 0.3  # ثانیه بین هر درخواست
PROGRESS_FILE = "enrich_progress.json"
FETCHED_DATA_FILE = "fetched_movies.json"
DIRECTORY = "."  # دایرکتوری حاوی فایل‌ها
# ======================================


def normalize(name):
    """نرمال‌سازی نام فیلم برای مقایسه"""
    if not name:
        return ''
    name = str(name).replace('\u200f', '').replace('\u200e', '').strip()
    name = re.sub(r'[\u2018\u2019\u201c\u201d\']', '', name)
    name = re.sub(r'[^\w\s\-]', ' ', name)
    name = re.sub(r'\s+', ' ', name).strip().lower()
    return name


def extract_name_year(path):
    """استخراج نام فیلم و سال از مسیر"""
    if not path:
        return None, None
    path = str(path)
    match = re.search(r'\\(\d+)\s*(.*)$', path)
    if match:
        name = match.group(2).strip()
        year = None
        # سال در انتهای نام
        ym = re.search(r'\s+(\d{4})\s*$', name)
        if ym:
            year = ym.group(1)
            name = name[:ym.start()].strip()
        # سال در پرانتز
        yp = re.search(r'\((\d{4})\)\s*$', name)
        if yp:
            year = yp.group(1)
            name = name[:yp.start()].strip()
        # حذف کاراکترهای RTL
        name = name.replace('\u200f', '').replace('\u200e', '').strip()
        return name, year
    return None, None


def clean_name_for_api(name):
    """پاک‌سازی نام برای ارسال به API"""
    clean = name.replace('\u2019', '').replace('\u2018', '')
    clean = clean.replace('\u201c', '').replace('\u201d', '')
    clean = re.sub(r'[^a-zA-Z0-9\s\-]', ' ', clean)
    clean = re.sub(r'\s+', ' ', clean).strip()
    return clean


def fetch_movie_info(name, year=None):
    """دریافت اطلاعات فیلم از OMDb API"""
    clean_name = clean_name_for_api(name)
    url = f"{BASE_URL}?t={quote(clean_name)}&apikey={API_KEY}"
    if year:
        url += f"&y={year}"
    
    try:
        response = requests.get(url, timeout=15)
        if response.status_code == 200:
            data = response.json()
            if data.get('Response') == 'True':
                return {
                    'Title': data.get('Title', ''),
                    'Year': data.get('Year', ''),
                    'Genre': data.get('Genre', ''),
                    'Director': data.get('Director', ''),
                    'Actors': data.get('Actors', ''),
                    'imdbRating': data.get('imdbRating', ''),
                    'Runtime': data.get('Runtime', ''),
                    'Country': data.get('Country', ''),
                    'Plot': data.get('Plot', ''),
                    'Poster': data.get('Poster', ''),
                    'Response': 'True'
                }
    except Exception as e:
        print(f"  ⚠ خطا در دریافت: {e}")
    
    # اگر با سال پیدا نشد، بدون سال امتحان کن
    if year:
        try:
            url2 = f"{BASE_URL}?t={quote(clean_name)}&apikey={API_KEY}"
            response2 = requests.get(url2, timeout=15)
            if response2.status_code == 200:
                data2 = response2.json()
                if data2.get('Response') == 'True':
                    return {
                        'Title': data2.get('Title', ''),
                        'Year': data2.get('Year', ''),
                        'Genre': data2.get('Genre', ''),
                        'Director': data2.get('Director', ''),
                        'Actors': data2.get('Actors', ''),
                        'imdbRating': data2.get('imdbRating', ''),
                        'Runtime': data2.get('Runtime', ''),
                        'Country': data2.get('Country', ''),
                        'Plot': data2.get('Plot', ''),
                        'Poster': data2.get('Poster', ''),
                        'Response': 'True'
                    }
        except:
            pass
    
    return None


def build_lookup(fetched_data):
    """ساخت دیکشنری جستجو از داده‌های دریافت شده"""
    lookup = {}
    for key, movie in fetched_data.items():
        if movie.get('Response') != 'True':
            continue
        nk = normalize(key)
        if nk:
            lookup[nk] = movie
        nt = normalize(movie.get('Title', ''))
        if nt:
            lookup[nt] = movie
    return lookup


def update_part_file(part_file, lookup):
    """به‌روزرسانی یک فایل قسمت"""
    wb = openpyxl.load_workbook(part_file)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_map = {h: i + 1 for i, h in enumerate(headers) if h}
    
    updated = 0
    for row in range(2, ws.max_row + 1):
        title_val = ws.cell(row=row, column=col_map.get('Title', 0)).value
        if title_val and str(title_val).strip():
            continue
        
        path = ws.cell(row=row, column=col_map.get('Movie Name', 0)).value
        name, year = extract_name_year(path)
        if not name:
            continue
        
        norm = normalize(name)
        movie_data = lookup.get(norm)
        
        if not movie_data:
            # جستجوی جزئی
            for k, d in lookup.items():
                if len(k) > 5 and (k in norm or norm in k):
                    movie_data = d
                    break
        
        if movie_data:
            ws.cell(row=row, column=col_map.get('Title', 0), value=movie_data.get('Title', ''))
            ws.cell(row=row, column=col_map.get('Original Title', 0), value=movie_data.get('Title', ''))
            ws.cell(row=row, column=col_map.get('Director', 0), value=movie_data.get('Director', ''))
            ws.cell(row=row, column=col_map.get('Cast', 0), value=movie_data.get('Actors', ''))
            ws.cell(row=row, column=col_map.get('Year', 0), value=movie_data.get('Year', ''))
            ws.cell(row=row, column=col_map.get('Genre', 0), value=movie_data.get('Genre', ''))
            ws.cell(row=row, column=col_map.get('Rating', 0), value=movie_data.get('imdbRating', ''))
            runtime = movie_data.get('Runtime', '')
            ws.cell(row=row, column=col_map.get('Runtime', 0), value=runtime.replace(' min', ''))
            ws.cell(row=row, column=col_map.get('Country', 0), value=movie_data.get('Country', ''))
            ws.cell(row=row, column=col_map.get('Synopsis', 0), value=movie_data.get('Plot', ''))
            ws.cell(row=row, column=col_map.get('Poster URL', 0), value=movie_data.get('Poster', ''))
            updated += 1
    
    wb.save(part_file)
    return updated


def merge_parts_to_main():
    """ادغام داده‌های قسمت‌ها به فایل اصلی"""
    print("\n📦 ادغام قسمت‌ها به فایل اصلی...")
    
    main_wb = openpyxl.load_workbook(os.path.join(DIRECTORY, '9_Movies.xlsx'))
    main_ws = main_wb.active
    headers = [cell.value for cell in main_ws[1]]
    col_map = {h: i + 1 for i, h in enumerate(headers) if h}
    
    total_merged = 0
    for part_num in range(1, 11):
        part_file = os.path.join(DIRECTORY, f'9_Movies_part{part_num}.xlsx')
        if not os.path.exists(part_file):
            continue
        
        part_wb = openpyxl.load_workbook(part_file)
        part_ws = part_wb.active
        part_headers = [cell.value for cell in part_ws[1]]
        part_col = {h: i + 1 for i, h in enumerate(part_headers) if h}
        
        # Read enriched rows from part
        enriched_rows = {}
        for row in range(2, part_ws.max_row + 1):
            path = part_ws.cell(row=row, column=part_col.get('Movie Name', 0)).value
            title = part_ws.cell(row=row, column=part_col.get('Title', 0)).value
            if title and str(title).strip() and path:
                enriched_rows[str(path)] = row
        
        # Match with main file
        for row in range(2, main_ws.max_row + 1):
            path = main_ws.cell(row=row, column=col_map.get('Movie Name', 0)).value
            title_val = main_ws.cell(row=row, column=col_map.get('Title', 0)).value
            
            if title_val and str(title_val).strip():
                continue
            
            if str(path) in enriched_rows:
                part_row = enriched_rows[str(path)]
                # Copy all enrichment columns
                for col_name in ['Title', 'Original Title', 'Director', 'Cast', 'Year', 'Genre', 'Rating', 'Runtime', 'Country', 'Synopsis', 'Poster URL']:
                    val = part_ws.cell(row=part_row, column=part_col.get(col_name, 0)).value
                    if val:
                        main_ws.cell(row=row, column=col_map.get(col_name, 0), value=val)
                total_merged += 1
        
        part_wb.close()
    
    main_wb.save(os.path.join(DIRECTORY, '9_Movies.xlsx'))
    print(f"  ✅ {total_merged} فیلم به فایل اصلی اضافه شد")


def enrich_all():
    """تکمیل اطلاعات همه فیلم‌ها"""
    print("=" * 60)
    print("🎬 شروع تکمیل اطلاعات فیلم‌های فایل 9")
    print("=" * 60)
    
    # بارگذاری داده‌های قبلی
    fetched_data = {}
    fetched_path = os.path.join(DIRECTORY, FETCHED_DATA_FILE)
    if os.path.exists(fetched_path):
        with open(fetched_path, 'r', encoding='utf-8') as f:
            fetched_data = json.load(f)
        print(f"📂 {len(fetched_data)} فیلم از قبل دریافت شده")
    
    # بارگذاری پیشرفت
    progress = {}
    progress_path = os.path.join(DIRECTORY, PROGRESS_FILE)
    if os.path.exists(progress_path):
        with open(progress_path, 'r', encoding='utf-8') as f:
            progress = json.load(f)
    
    # پردازش هر قسمت
    for part_num in range(1, 11):
        part_file = os.path.join(DIRECTORY, f'9_Movies_part{part_num}.xlsx')
        if not os.path.exists(part_file):
            print(f"\n⚠ فایل {part_file} پیدا نشد")
            continue
        
        print(f"\n{'─' * 40}")
        print(f"📁 قسمت {part_num}/10")
        print(f"{'─' * 40}")
        
        wb = openpyxl.load_workbook(part_file)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        col_map = {h: i + 1 for i, h in enumerate(headers) if h}
        
        total = ws.max_row - 1
        enriched = sum(1 for r in range(2, ws.max_row + 1)
                       if ws.cell(row=r, column=col_map.get('Title', 0)).value)
        remaining = total - enriched
        print(f"  📊 {enriched}/{total} تکمیل شده ({remaining} باقیمانده)")
        
        if remaining == 0:
            print("  ✅ این قسمت کامل است!")
            wb.close()
            continue
        
        # بررسی نقطه شروع
        start_from = progress.get(f'part_{part_num}', {}).get('last_row', 2)
        
        processed = 0
        found = 0
        not_found = 0
        errors = 0
        
        for row in range(start_from, ws.max_row + 1):
            title_val = ws.cell(row=row, column=col_map.get('Title', 0)).value
            if title_val and str(title_val).strip():
                continue
            
            path = ws.cell(row=row, column=col_map.get('Movie Name', 0)).value
            name, year = extract_name_year(path)
            if not name:
                continue
            
            # نمایش پیشرفت
            progress_pct = (row / ws.max_row) * 100
            year_str = f" ({year})" if year else ""
            print(f"  [{progress_pct:.0f}%] {name}{year_str}", end=" ")
            
            # دریافت اطلاعات
            movie_data = fetch_movie_info(name, year)
            
            if movie_data:
                # ذخیره در داده‌های دریافت شده
                fetched_data[name] = movie_data
                
                # به‌روزرسانی اکسل
                ws.cell(row=row, column=col_map.get('Title', 0), value=movie_data.get('Title', ''))
                ws.cell(row=row, column=col_map.get('Original Title', 0), value=movie_data.get('Title', ''))
                ws.cell(row=row, column=col_map.get('Director', 0), value=movie_data.get('Director', ''))
                ws.cell(row=row, column=col_map.get('Cast', 0), value=movie_data.get('Actors', ''))
                ws.cell(row=row, column=col_map.get('Year', 0), value=movie_data.get('Year', ''))
                ws.cell(row=row, column=col_map.get('Genre', 0), value=movie_data.get('Genre', ''))
                ws.cell(row=row, column=col_map.get('Rating', 0), value=movie_data.get('imdbRating', ''))
                runtime = movie_data.get('Runtime', '')
                ws.cell(row=row, column=col_map.get('Runtime', 0), value=runtime.replace(' min', ''))
                ws.cell(row=row, column=col_map.get('Country', 0), value=movie_data.get('Country', ''))
                ws.cell(row=row, column=col_map.get('Synopsis', 0), value=movie_data.get('Plot', ''))
                ws.cell(row=row, column=col_map.get('Poster URL', 0), value=movie_data.get('Poster', ''))
                
                found += 1
                print("✓")
            else:
                not_found += 1
                print("✗ (پیدا نشد)")
            
            processed += 1
            
            # ذخیره هر 10 فیلم
            if processed % 10 == 0:
                wb.save(part_file)
                with open(fetched_path, 'w', encoding='utf-8') as f:
                    json.dump(fetched_data, f, ensure_ascii=False, indent=2)
                with open(progress_path, 'w', encoding='utf-8') as f:
                    json.dump({
                        **progress,
                        f'part_{part_num}': {'last_row': row + 1, 'found': found, 'not_found': not_found}
                    }, f, ensure_ascii=False, indent=2)
            
            # احترام به rate limit
            time.sleep(RATE_LIMIT_DELAY)
        
        # ذخیره نهایی
        wb.save(part_file)
        wb.close()
        
        with open(fetched_path, 'w', encoding='utf-8') as f:
            json.dump(fetched_data, f, ensure_ascii=False, indent=2)
        
        print(f"\n  📈 نتایج قسمت {part_num}:")
        print(f"     ✅ پیدا شده: {found}")
        print(f"     ❌ پیدا نشده: {not_found}")
        print(f"     📊 کل دریافت شده: {len(fetched_data)}")
    
    # ادغام به فایل اصلی
    merge_parts_to_main()
    
    # گزارش نهایی
    print(f"\n{'=' * 60}")
    print(f"✅ تکمیل شد!")
    print(f"{'=' * 60}")
    
    total_enriched = 0
    total_movies = 0
    for part_num in range(1, 11):
        part_file = os.path.join(DIRECTORY, f'9_Movies_part{part_num}.xlsx')
        if not os.path.exists(part_file):
            continue
        wb = openpyxl.load_workbook(part_file)
        ws = wb.active
        part_total = ws.max_row - 1
        part_enriched = sum(1 for r in range(2, ws.max_row + 1)
                           if ws.cell(row=r, column=col_map.get('Title', 0)).value if 'Title' in col_map)
        total_enriched += part_enriched
        total_movies += part_total
        wb.close()
        print(f"  قسمت {part_num}: {part_enriched}/{part_total}")
    
    print(f"\n  📊 مجموع: {total_enriched}/{total_movies} فیلم تکمیل شده")
    print(f"  📊 درصد تکمیل: {(total_enriched/total_movies*100):.1f}%")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    enrich_all()
