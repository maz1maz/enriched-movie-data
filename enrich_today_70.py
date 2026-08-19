#!/usr/bin/env python3
"""
تکمیل فایل Movies_Catalog_Today_70.xlsx
اضافه کردن ستون‌های جدید از TMDb API
"""

import openpyxl
import json
import os
import time
import urllib.request
import urllib.parse
import ssl

EXCEL_FILE = 'Movies_Catalog_Today_70.xlsx'
TMDB_API_KEY = '6e0aff711545aa88cd101d88b235d320'
TMDB_BASE_URL = 'https://api.themoviedb.org/3'
DELAY = 2.5

ssl_ctx = ssl.create_default_context()
ssl_ctx.check_hostname = False
ssl_ctx.verify_mode = ssl.CERT_NONE

def tmdb_request(url, params=None):
    if params is None:
        params = {}
    params['api_key'] = TMDB_API_KEY
    params['language'] = 'en-US'
    query_string = urllib.parse.urlencode(params)
    full_url = f"{TMDB_BASE_URL}{url}?{query_string}"
    
    for attempt in range(3):
        try:
            req = urllib.request.Request(full_url, headers={'User-Agent': 'MovieCatalog/1.0', 'Accept': 'application/json'})
            response = urllib.request.urlopen(req, timeout=15, context=ssl_ctx)
            return json.loads(response.read().decode())
        except urllib.error.HTTPError as e:
            if e.code == 429:
                time.sleep((attempt + 1) * 5)
                continue
            return None
        except:
            if attempt < 2:
                time.sleep(2)
            else:
                return None
    return None

def get_genres_list():
    data = tmdb_request('/genre/movie/list')
    if data and 'genres' in data:
        return {g['id']: g['name'] for g in data['genres']}
    return {}

def search_movie(title, year=None):
    params = {'query': title}
    if year:
        params['year'] = year
    data = tmdb_request('/search/movie', params)
    if data and 'results' in data and len(data['results']) > 0:
        return data['results'][0]
    return None

def get_movie_details(movie_id):
    return tmdb_request(f'/movie/{movie_id}')

def get_movie_credits(movie_id):
    return tmdb_request(f'/movie/{movie_id}/credits')

def find_column_index(ws, possible_names):
    headers = [cell.value for cell in ws[1]]
    for i, h in enumerate(headers, 1):
        if h:
            for name in possible_names:
                if name.lower() in str(h).lower():
                    return i, h
    return None, None

def update_excel():
    if not os.path.exists(EXCEL_FILE):
        print(f" File {EXCEL_FILE} not found!")
        return
    
    print(f"Reading {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    
    new_columns = {
        'Genre': 'Genre',
        'Country': 'Country',
        'Cast': 'Cast',
        'Synopsis': 'Synopsis',
        'Poster': 'Poster URL',
        'MPA Rating': 'MPA Rating',
        'Language': 'Language',
        'Box Office': 'Box Office',
        'Tagline': 'Tagline',
        'Budget': 'Budget',
        'Release Date': 'Release Date',
        'Status': 'Status',
        'Popularity': 'Popularity',
        'Homepage': 'Homepage',
        'Production Companies': 'Production Companies'
    }
    
    for col_name in new_columns:
        found = False
        for h in headers:
            if h and col_name.lower() in str(h).lower():
                found = True
                break
        if not found:
            last_col = len(headers) + 1
            ws.cell(row=1, column=last_col, value=new_columns[col_name])
            headers.append(new_columns[col_name])
            print(f"  Added '{new_columns[col_name]}' column")
    
    print("\nLoading genres...")
    genres_map = get_genres_list()
    print(f"  {len(genres_map)} genres loaded")
    
    title_col, _ = find_column_index(ws, ['Title', 'title'])
    year_col, _ = find_column_index(ws, ['Year', 'year'])
    
    total_movies = ws.max_row - 1
    print(f"\nTotal: {total_movies} films\n")
    
    updated = 0
    skipped = 0
    errors = 0
    
    for row in range(2, ws.max_row + 1):
        title = ws.cell(row=row, column=title_col).value
        if not title:
            continue
        
        genre_col_idx, _ = find_column_index(ws, ['Genre'])
        if genre_col_idx:
            genre_val = ws.cell(row=row, column=genre_col_idx).value
            if genre_val and str(genre_val).strip():
                skipped += 1
                continue
        
        year = ws.cell(row=row, column=year_col).value if year_col else None
        year_str = str(year).replace('.0', '') if year else None
        
        print(f" [{row-1}/{total_movies}] {title[:40]} ({year_str or '?'})...", end=' ')
        
        search_result = search_movie(title, year_str)
        
        if search_result:
            movie_id = search_result.get('id')
            details = get_movie_details(movie_id)
            
            if details:
                genre_ids = details.get('genres', [])
                genres = ', '.join([genres_map.get(g['id'], g['name']) for g in genre_ids])
                countries = details.get('production_countries', [])
                country = ', '.join([c['name'] for c in countries[:2]])
                synopsis = details.get('overview', '')
                poster_path = details.get('poster_path', '')
                poster_url = f"https://image.tmdb.org/t/p/w500{poster_path}" if poster_path else ''
                
                credits_data = get_movie_credits(movie_id)
                cast = ''
                if credits_data and 'cast' in credits_data:
                    cast = ', '.join([m['name'] for m in credits_data['cast'][:5]])
                
                for col_name in ['Genre', 'Country', 'Cast', 'Synopsis', 'Poster URL', 'MPA Rating', 'Language', 'Box Office', 'Tagline', 'Budget', 'Release Date', 'Status', 'Popularity', 'Homepage', 'Production Companies']:
                    col_idx, _ = find_column_index(ws, [col_name])
                    if col_idx:
                        value = {
                            'Genre': genres,
                            'Country': country,
                            'Cast': cast,
                            'Synopsis': synopsis,
                            'Poster URL': poster_url,
                            'MPA Rating': details.get('certification', ''),
                            'Language': ', '.join([l['name'] for l in details.get('spoken_languages', [])[:3]]),
                            'Box Office': details.get('revenue', ''),
                            'Tagline': details.get('tagline', ''),
                            'Budget': details.get('budget', ''),
                            'Release Date': details.get('release_date', ''),
                            'Status': details.get('status', ''),
                            'Popularity': details.get('popularity', ''),
                            'Homepage': details.get('homepage', ''),
                            'Production Companies': ', '.join([c['name'] for c in details.get('production_companies', [])[:3]])
                        }.get(col_name, '')
                        ws.cell(row=row, column=col_idx, value=value)
                
                updated += 1
                print(f"OK")
            else:
                errors += 1
                print("ERR")
        else:
            errors += 1
            print("NOT FOUND")
        
        if (updated + skipped + errors) % 10 == 0:
            wb.save(EXCEL_FILE)
        
        time.sleep(DELAY)
    
    wb.save(EXCEL_FILE)
    
    print(f"\n{'='*60}")
    print(f"Done!")
    print(f"Updated: {updated}")
    print(f"Skipped: {skipped}")
    print(f"Errors: {errors}")
    print(f"Total: {total_movies}")
    print(f"{'='*60}")

if __name__ == '__main__':
    print("=" * 60)
    print("Movies_Catalog_Today_70.xlsx Enricher")
    print("=" * 60)
    update_excel()
