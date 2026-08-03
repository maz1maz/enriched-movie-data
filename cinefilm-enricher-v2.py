#!/usr/bin/env python3
"""
Cinefilm Archive Enricher v2
==============================
نسخه بهبودیافته و ترکیبی از بهترین ویژگی‌های دو نسخه قبلی

ویژگی‌ها:
  ✅ ترجمه خودکار عناوین فارسی/غیرانگلیسی به انگلیسی
  ✅ جستجوی IMDb ID از طریق IMDb Suggestion API (دقت بالاتر)
  ✅ پشتیبانی فیلم (OMDb) + سریال (TVMaze)
  ✅ سیستم کش پیشرفته (resume بعد از قطعی)
  ✅ تشخیص خودکار نوع فایل (فیلم یا سریال)
  ✅ پشتیبانی CSV + Excel + انکدینگ‌های مختلف
  ✅ خروجی Missing برای اصلاح دستی
  ✅ ستون‌های کامل: Shelf, Row, Director, Cast, Year, Genre, Rating, Runtime,
     Country, Synopsis, Poster URL, Original Title, Total Size, Folder Paths
  ✅ اصلاح خودکار عناوین مشکل‌دار
  ✅ نوار پیشرفت (progress bar)
  ✅ مدیریت Rate Limit هوشمند
  ✅ گزارش نهایی آمار
  ✅ دور زدن خودکار مشکلات SSL (با ۳ روش fallback: requests, httpx, urllib)
  ✅ استخراج خودکار نام فیلم از مسیر فایل (مثل H:\Movies\...)
  ✅ وارد/صادر کردن کش از/به فایل JSON خارجی
  ✅ ادامه کار از کش قبلی بدون نیاز به شروع مجدد

Requirements:
    pip install requests pandas openpyxl deep-translator chardet
    # اختیاری (برای دور زدن بهتر SSL):
    pip install httpx

Usage:
    python cinefilm-enricher-v2.py --input movies.csv
    python cinefilm-enricher-v2.py --input series.xlsx --type series
    python cinefilm-enricher-v2.py --input data.csv --type auto --output results/
    python cinefilm-enricher-v2.py --input data.csv --template film-archive-template.xlsx
    python cinefilm-enricher-v2.py --input data.csv --delay 0.5 --no-translate
    python cinefilm-enricher-v2.py --input data.csv --import-cache existing_cache.json
    python cinefilm-enricher-v2.py --input data.csv --export-cache new_cache.json
"""

import argparse
import csv
import json
import os
import re
import sys
import time
import html
import urllib.parse
from pathlib import Path
from collections import defaultdict
from copy import copy as copy_style

import requests
import ssl
import urllib.request
import urllib.error

# تلاش برای httpx (بهتر از requests برای SSL)
try:
    import httpx
    HAS_HTTPX = True
except ImportError:
    HAS_HTTPX = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import chardet
    HAS_CHARDET = True
except ImportError:
    HAS_CHARDET = False

try:
    from deep_translator import GoogleTranslator
    HAS_TRANSLATOR = True
except ImportError:
    HAS_TRANSLATOR = False

# =============================================================================
# CONFIG
# =============================================================================
OMDB_API_KEY = "fe53f97e"
OMDB_URL = "https://www.omdbapi.com/"
IMDB_SUGGESTION_URL = "https://v3.sg.media-imdb.com/suggestion/{}/{}.json"
TVMAZE_SEARCH_URL = "https://api.tvmaze.com/search/shows"
TVMAZE_SHOW_URL = "https://api.tvmaze.com/shows/{}"
TVMAZE_CAST_URL = "https://api.tvmaze.com/shows/{}/cast"
TVMAZE_CREW_URL = "https://api.tvmaze.com/shows/{}/crew"

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

# اگر این فایل کنار برنامه/ورودی باشد، خروجی Excel نهایی با همین قالب ساخته می‌شود.
DEFAULT_TEMPLATE_FILENAME = "film-archive-template.xlsx"

# ستون‌های نهایی خروجی
MOVIE_COLUMNS = [
    "Title", "Shelf", "Row", "Director", "Cast", "Year",
    "Genre", "Rating", "Runtime", "Country", "Synopsis",
    "Poster URL", "Original Title", "Total Size", "Folder Paths"
]

SERIES_COLUMNS = [
    "Title", "Shelf", "Row", "Director", "Cast", "Year",
    "Genre", "Rating", "Runtime", "Country", "Synopsis",
    "Poster URL", "Original Title", "Seasons Available",
    "Total Size", "Folder Paths"
]

# اصلاحات دستی برای عناوین مشکل‌دار (TVMaze ID)
TVMAZE_CORRECTIONS = {
    "money heist": 27436,
    "la casa de papel": 27436,
    "protector": 36807,
    "the protector": 36807,
    "anne": 12989,
    "anne with an e": 12989,
    "24": 406,
    "the bridge": 1910,
    "dark": 13177,
}

# اصلاحات دستی برای عناوین مشکل‌دار (OMDb جستجوی جایگزین)
OMDB_TITLE_CORRECTIONS = {
    "money heist": "La Casa de Papel",
}

# =============================================================================
# CACHE SYSTEM
# =============================================================================
class Cache:
    """سیستم کش ساده و مؤثر — درخواست‌های تکراری رو حذف می‌کنه"""

    def __init__(self, cache_path):
        self.path = Path(cache_path)
        self.data = {}
        self.hits = 0
        self.misses = 0
        self._dirty = False
        self._load()

    def _load(self):
        if self.path.exists():
            try:
                with open(self.path, "r", encoding="utf-8") as f:
                    self.data = json.load(f)
                print(f"  📦 کش بارگذاری شد: {len(self.data)} آیتم ({self.path.name})")
            except (json.JSONDecodeError, IOError):
                self.data = {}

    def save(self):
        if self._dirty:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            with open(self.path, "w", encoding="utf-8") as f:
                json.dump(self.data, f, ensure_ascii=False, indent=1)
            self._dirty = False

    def get(self, key):
        key = str(key).strip().lower()
        if key in self.data:
            self.hits += 1
            return self.data[key]
        self.misses += 1
        return None

    def set(self, key, value):
        self.data[str(key).strip().lower()] = value
        self._dirty = True

    def stats(self):
        total = self.hits + self.misses
        if total == 0:
            return "بدون درخواست"
        pct = (self.hits / total) * 100
        return f"✅ {self.hits} کش‌هیت / ❌ {self.misses} درخواست جدید ({pct:.0f}% صرفه‌جویی)"

    def import_from(self, external_path):
        """وارد کردن داده‌های کش از فایل JSON خارجی"""
        ext_path = Path(external_path)
        if not ext_path.exists():
            print(f"  ⚠️ فایل کش یافت نشد: {ext_path}")
            return 0
        try:
            with open(ext_path, "r", encoding="utf-8") as f:
                ext_data = json.load(f)
            
            imported = 0
            for key, value in ext_data.items():
                # تبدیل کلید به فرمت استاندارد کش (movie:TITLE)
                if ":" not in key:
                    std_key = f"movie:{key}"
                else:
                    std_key = key
                std_key = std_key.strip().lower()
                if std_key not in self.data:
                    self.data[std_key] = value
                    imported += 1
                    self._dirty = True
            
            return imported
        except (json.JSONDecodeError, IOError) as e:
            print(f"  ⚠️ خطا در وارد کردن کش: {e}")
            return 0

    def export_to(self, export_path):
        """خروجی کش به فایل JSON"""
        export_path = Path(export_path)
        export_path.parent.mkdir(parents=True, exist_ok=True)
        with open(export_path, "w", encoding="utf-8") as f:
            json.dump(self.data, f, ensure_ascii=False, indent=1)
        print(f"  💾 کش صادر شد: {export_path}")


# =============================================================================
# TRANSLATION
# =============================================================================
def is_pure_ascii(s):
    try:
        s.encode("ascii")
        return True
    except UnicodeEncodeError:
        return False


def translate_to_english(query):
    """ترجمه خودکار عنوان غیرانگلیسی به انگلیسی"""
    if not HAS_TRANSLATOR:
        return query
    if is_pure_ascii(query):
        return query
    try:
        translated = GoogleTranslator(source="auto", target="en").translate(query)
        if translated and translated.strip():
            return translated.strip()
    except Exception:
        pass
    return query


# =============================================================================
# TEXT CLEANING
# =============================================================================
def strip_html(text):
    if not text:
        return ""
    text = re.sub(r"<[^>]+>", "", text)
    return html.unescape(text).strip()


def extract_name_from_path(path_str):
    """استخراج نام فیلم/سریال از مسیر فایل ویندوز (مثل H:\\Movies\\100 metros)"""
    if not path_str:
        return path_str
    path_str = str(path_str).strip()
    # تبدیل بک‌اسلش به اسلش
    normalized = path_str.replace("\\", "/")
    # گرفتن آخرین بخش مسیر
    parts = normalized.split("/")
    name = parts[-1] if parts else path_str
    # حذف پرانتز ناتمام مثل "(2016"
    if "(" in name and ")" not in name:
        name = name[:name.rfind("(")].strip()
    # حذف اعداد ترتیبی (مثل "1. Movie Name")
    name = re.sub(r"^\d+[\.\)]\s*", "", name)
    return name.strip()


def is_file_path(value):
    """تشخیص اینکه آیا مقدار یک مسیر فایل است"""
    s = str(value).strip()
    # الگوهای مسیر ویندوز
    if re.match(r"^[A-Z]:[/\\\\]", s):
        return True
    if "\\" in s and re.search(r"[/\\\\][^/\\\\]+$", s):
        return True
    return False


def clean_title(title):
    """تمیز کردن عنوان فیلم/سریال"""
    t = str(title).strip()
    t = re.sub(r"\s+", " ", t)
    # حذف تگ‌های رایج
    t = re.sub(r"\s*\(TV Movie.*$", "", t, flags=re.I)
    t = re.sub(r"\s*\(Video.*$", "", t, flags=re.I)
    t = re.sub(r"\s*\(Short.*$", "", t, flags=re.I)
    # حذف کاراکترهای عجیب
    t = t.replace("\ufffd", "").replace("\xb3", "3").strip()
    return t


def clean_runtime(runtime_str):
    """استخراج عدد از رشته Runtime"""
    if not runtime_str or runtime_str == "N/A":
        return ""
    match = re.search(r"(\d+)", str(runtime_str))
    if match:
        return int(match.group(1))
    return runtime_str


# =============================================================================
# HTTP REQUEST WITH RETRY + SSL FALLBACK
# =============================================================================

class HTTPResponse:
    """Wrapper برای یکپارچه‌سازی پاسخ‌ها از روش‌های مختلف"""
    def __init__(self, status_code, text, json_data=None):
        self.status_code = status_code
        self.text = text
        self._json = json_data

    def json(self):
        if self._json is not None:
            return self._json
        return json.loads(self.text)


def _try_requests(url, params=None, headers=None, timeout=10):
    """تلاش ۱: استفاده از requests با SSL context سفارشی"""
    try:
        session = requests.Session()
        # تنظیم SSL context برای دور زدن مشکلات
        adapter = requests.adapters.HTTPAdapter(max_retries=0)
        session.mount('https://', adapter)
        res = session.get(url, params=params, headers=headers, timeout=timeout,
                          verify=False)  # غیرفعال‌سازی verify برای محیط‌های محدود
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        return HTTPResponse(res.status_code, res.text, None)
    except Exception:
        return None


def _try_urllib(url, timeout=10):
    """تلاش ۲: استفاده از urllib با SSL context بدون verify"""
    try:
        ctx = ssl._create_unverified_context()
        req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
        with urllib.request.urlopen(req, timeout=timeout, context=ctx) as response:
            data = response.read().decode("utf-8")
            return HTTPResponse(response.status, data, None)
    except Exception:
        return None


def _try_httpx(url, params=None, headers=None, timeout=10):
    """تلاش ۳: استفاده از httpx"""
    if not HAS_HTTPX:
        return None
    try:
        with httpx.Client(verify=False, timeout=timeout) as client:
            res = client.get(url, params=params, headers=headers)
            return HTTPResponse(res.status_code, res.text, None)
    except Exception:
        return None


def _build_url_with_params(url, params=None):
    """ساخت URL با پارامترها"""
    if params:
        query = urllib.parse.urlencode(params)
        separator = "&" if "?" in url else "?"
        return f"{url}{separator}{query}"
    return url


def http_get_with_retry(url, params=None, headers=None, timeout=10, max_retries=3):
    """
    درخواست HTTP با چند روش (fallback) و retry
    
    روش‌ها به ترتیب:
      1. requests (با verify=False)
      2. httpx (در صورت نصب)
      3. urllib (با SSL context بدون verify)
    
    هر روش تا max_retries بار تلاش می‌کند.
    """
    full_url = _build_url_with_params(url, params)
    
    for attempt in range(max_retries):
        # روش ۱: requests
        res = _try_requests(url, params=params, headers=headers, timeout=timeout)
        if res:
            if res.status_code == 429:
                wait_time = (2 ** attempt) + 1
                time.sleep(wait_time)
                continue
            return res
        
        # روش ۲: httpx
        res = _try_httpx(url, params=params, headers=headers, timeout=timeout)
        if res:
            if res.status_code == 429:
                wait_time = (2 ** attempt) + 1
                time.sleep(wait_time)
                continue
            return res
        
        # روش ۳: urllib
        res = _try_urllib(full_url, timeout=timeout)
        if res:
            return res
        
        if attempt < max_retries - 1:
            wait_time = (2 ** attempt) + 0.5
            time.sleep(wait_time)
    
    return None


# =============================================================================
# IMDb SUGGESTION API
# =============================================================================
def get_imdb_id_from_suggestion(query):
    """
    پیدا کردن IMDb ID از طریق سرویس Suggestion IMDb
    این روش دقیق‌تر از جستجوی مستقیم عنوان در OMDb هست
    """
    try:
        clean_q = query.strip().lower()
        first_char = clean_q[0]
        encoded = urllib.parse.quote(clean_q)
        url = IMDB_SUGGESTION_URL.format(first_char, encoded)
        headers = {"User-Agent": USER_AGENT}
        res = http_get_with_retry(url, headers=headers)
        if res and res.status_code == 200:
            data = res.json()
            if "d" in data and data["d"]:
                # اولویت با فیلم یا سریال
                for item in data["d"]:
                    qid = item.get("qid", "")
                    if qid in ("movie", "tvSeries", "tvMiniSeries"):
                        return item.get("id"), qid
                # اگه هیچکدوم فیلم/سریال نبود، اولی رو برگردون
                return data["d"][0].get("id"), data["d"][0].get("qid", "")
    except Exception:
        pass
    return None, None


# =============================================================================
# OMDb API (فیلم)
# =============================================================================
def fetch_movie_details(title, cache, translate=True, verbose=True):
    """دریافت اطلاعات کامل فیلم از OMDb"""
    # بررسی کش
    cached = cache.get(f"movie:{title}")
    if cached is not None:
        return cached if cached != {} else None

    search_title = title
    if translate and not is_pure_ascii(title):
        if verbose:
            print(f"  🌐 ترجمه '{title}'...", end=" ", flush=True)
        search_title = translate_to_english(title)
        if verbose and search_title != title:
            print(f"→ '{search_title}'")
        elif verbose:
            print("(بدون تغییر)")

    # اصلاح عنوان
    search_lower = search_title.lower().strip()
    if search_lower in OMDB_TITLE_CORRECTIONS:
        search_title = OMDB_TITLE_CORRECTIONS[search_lower]

    omdb_data = None

    # مرحله ۱: IMDb Suggestion → IMDb ID → OMDb
    imdb_id, qid = get_imdb_id_from_suggestion(search_title)
    if imdb_id:
        res = http_get_with_retry(OMDB_URL, params={
            "apikey": OMDB_API_KEY,
            "i": imdb_id,
            "plot": "full",
        })
        if res and res.status_code == 200:
            try:
                data = res.json()
                if data.get("Response") == "True":
                    omdb_data = data
            except Exception:
                pass

    # مرحله ۲: Fallback — جستجوی مستقیم عنوان
    if not omdb_data:
        res = http_get_with_retry(OMDB_URL, params={
            "apikey": OMDB_API_KEY,
            "t": search_title,
            "plot": "full",
        })
        if res and res.status_code == 200:
            try:
                data = res.json()
                if data.get("Response") == "True":
                    omdb_data = data
            except Exception:
                pass

    # ذخیره در کش (حتی None برای جلوگیری از درخواست تکراری)
    cache.set(f"movie:{title}", omdb_data if omdb_data else {})
    return omdb_data


# =============================================================================
# TVMaze API (سریال)
# =============================================================================
def fetch_series_details(title, cache, translate=True, verbose=True):
    """دریافت اطلاعات کامل سریال از TVMaze"""
    # بررسی کش
    cached = cache.get(f"series:{title}")
    if cached is not None:
        return cached if cached != {} else None

    search_title = title
    if translate and not is_pure_ascii(title):
        if verbose:
            print(f"  🌐 ترجمه '{title}'...", end=" ", flush=True)
        search_title = translate_to_english(title)
        if verbose and search_title != title:
            print(f"→ '{search_title}'")
        elif verbose:
            print("(بدون تغییر)")

    show_id = None
    search_lower = search_title.lower().strip()

    # بررسی اصلاحات دستی
    if search_lower in TVMAZE_CORRECTIONS:
        show_id = TVMAZE_CORRECTIONS[search_lower]

    # جستجو در TVMaze
    if not show_id:
        res = http_get_with_retry(TVMAZE_SEARCH_URL, params={"q": search_title})
        if res and res.status_code == 200:
            try:
                results = res.json()
                if results:
                    # بهترین نتیجه
                    best = results[0]
                    score = best.get("score", 0)
                    show = best.get("show", {})
                    # اگه امتیاز خیلی کمه، مطمئن نیستیم
                    if score > 0.5 or show:
                        show_id = show.get("id")
            except Exception:
                pass

    if not show_id:
        cache.set(f"series:{title}", {})
        return None

    # دریافت جزئیات سریال
    details = {}
    res = http_get_with_retry(TVMAZE_SHOW_URL.format(show_id))
    if res and res.status_code == 200:
        try:
            show = res.json()
            details["Title"] = show.get("name", "")
            details["Original Title"] = show.get("name", "")

            # سال
            premiered = show.get("premiered", "")
            if premiered:
                details["Year"] = premiered[:4]

            # ژانر
            genres = show.get("genres", [])
            details["Genre"] = ", ".join(genres) if genres else ""

            # رتبه
            rating = show.get("rating", {})
            details["Rating"] = str(rating.get("average", "")) if rating else ""

            # Runtime
            details["Runtime"] = show.get("runtime", "") or ""

            # خلاصه
            details["Synopsis"] = strip_html(show.get("summary", ""))

            # کشور
            network = show.get("network", {})
            country = network.get("country", {}) if network else {}
            details["Country"] = country.get("name", "")

            # پوستر
            image = show.get("image", {})
            details["Poster URL"] = image.get("original", "") or image.get("medium", "") or ""

            # تعداد فصل‌ها
            details["_seasons"] = show.get("seasons", [])
            details["_show_id"] = show_id
        except Exception:
            pass

    # دریافت بازیگران و کارگردان
    res = http_get_with_retry(TVMAZE_CAST_URL.format(show_id))
    if res and res.status_code == 200:
        try:
            cast_data = res.json()
            cast_names = [p.get("person", {}).get("name", "") for p in cast_data[:8]]
            details["Cast"] = ", ".join(filter(None, cast_names))
        except Exception:
            pass

    # کارگردان از crew — TVMaze crew API محدود است
    details["Director"] = ""

    # تعداد فصل‌ها
    seasons = details.pop("_seasons", [])
    details["Seasons Available"] = len(seasons) if seasons else ""

    cache.set(f"series:{title}", details if details else {})
    return details if details else None


# =============================================================================
# FILE TYPE AUTO-DETECTION
# =============================================================================
def detect_file_type(file_path, df=None):
    """تشخیص خودکار نوع فایل: movie یا series"""
    name = str(file_path).lower()

    # از نام فایل
    if "series" in name or "serial" in name or "tv" in name:
        return "series"
    if "movie" in name or "film" in name:
        return "movie"

    # از محتوای فایل
    if df is not None:
        all_text = " ".join(str(x) for x in df.head(20).values.flatten() if x)
        lower_text = all_text.lower()
        series_keywords = ["season", "episode", "s01", "s02", "series", "tv series", "فصل", "قسمت"]
        movie_keywords = ["movie", "film", "director", "فیلم"]

        series_score = sum(1 for k in series_keywords if k in lower_text)
        movie_score = sum(1 for k in movie_keywords if k in lower_text)

        if series_score > movie_score:
            return "series"
        if movie_score > series_score:
            return "movie"

    # پیش‌فرض
    return "movie"



# =============================================================================
# COLUMN NORMALIZATION + EXCEL TEMPLATE OUTPUT
# =============================================================================
def column_key(name):
    """کلید ساده برای مقایسه نام ستون‌ها (فارسی/انگلیسی/فاصله‌ها را تحمل می‌کند)."""
    return re.sub(r"[^a-z0-9]+", "", str(name or "").lower())


COLUMN_ALIASES = {
    # title
    "title": "Title",
    "name": "Title",
    "movietitle": "Title",
    "filmtitle": "Title",
    # template fields
    "shelf": "Shelf",
    "row": "Row",
    "director": "Director",
    "directors": "Director",
    "cast": "Cast",
    "actors": "Cast",
    "actor": "Cast",
    "year": "Year",
    "releaseyear": "Year",
    "genre": "Genre",
    "genres": "Genre",
    "rating": "Rating",
    "imdbrating": "Rating",
    "imdbscore": "Rating",
    "runtime": "Runtime",
    "duration": "Runtime",
    "country": "Country",
    "countries": "Country",
    "synopsis": "Synopsis",
    "plot": "Synopsis",
    "summary": "Synopsis",
    "story": "Synopsis",
    "posterurl": "Poster URL",
    "poster": "Poster URL",
    "posterlink": "Poster URL",
    "posterimage": "Poster URL",
    "originaltitle": "Original Title",
    "totalsize": "Total Size",
    "folderpaths": "Folder Paths",
    "seasonsavailable": "Seasons Available",
}


def canonical_column_name(name):
    """نام استاندارد ستون را برمی‌گرداند، اگر ستون شناخته‌شده باشد."""
    key = column_key(name)
    if key in COLUMN_ALIASES:
        return COLUMN_ALIASES[key]

    # هدرهایی مثل «عنوان فیلم (Title)» یا mojibake دارای (Title) را هم تشخیص بده.
    text = str(name or "")
    paren = re.search(r"\(([^)]+)\)", text)
    if paren:
        pkey = column_key(paren.group(1))
        if pkey in COLUMN_ALIASES:
            return COLUMN_ALIASES[pkey]

    return str(name).strip()


def dataframe_column_lookup(df):
    """مپ کلید ستون → نام واقعی ستون در DataFrame برای lookup انعطاف‌پذیر."""
    lookup = {}
    for col in df.columns:
        clean = str(col).strip()
        lookup.setdefault(column_key(clean), col)
        canonical = canonical_column_name(clean)
        lookup.setdefault(column_key(canonical), col)
    return lookup


def normalize_output_value(header, value):
    """مقدار خروجی را برای قالب نهایی تمیز می‌کند."""
    if value is None:
        return ""
    if HAS_PANDAS:
        try:
            if pd.isna(value):
                return ""
        except Exception:
            pass
    if str(value).strip().upper() == "N/A":
        return ""
    if canonical_column_name(header) == "Runtime":
        return clean_runtime(value)
    return value


def build_dataframe_for_headers(df, headers):
    """DataFrame خروجی را دقیقاً با ستون‌های قالب Excel می‌سازد."""
    if not (HAS_PANDAS and isinstance(df, pd.DataFrame)):
        return None

    col_lookup = dataframe_column_lookup(df)
    rows = []
    for _, row in df.iterrows():
        out = {}
        for header in headers:
            canonical = canonical_column_name(header)
            value = ""
            for candidate in (header, canonical):
                col = col_lookup.get(column_key(candidate))
                if col is not None:
                    value = row.get(col, "")
                    break
            out[header] = normalize_output_value(header, value)
        rows.append(out)
    return pd.DataFrame(rows, columns=headers)


def resolve_template_path(template_path=None, input_path=None, use_template=True):
    """پیدا کردن قالب Excel. اگر مسیر داده نشود، film-archive-template.xlsx خودکار پیدا می‌شود."""
    if not use_template:
        return None

    candidates = []
    if template_path:
        candidates.append(Path(template_path))
    else:
        if input_path:
            candidates.append(Path(input_path).parent / DEFAULT_TEMPLATE_FILENAME)
        candidates.append(Path.cwd() / DEFAULT_TEMPLATE_FILENAME)
        candidates.append(Path(__file__).resolve().parent / DEFAULT_TEMPLATE_FILENAME)

    seen = set()
    for candidate in candidates:
        candidate = candidate.expanduser()
        key = str(candidate.resolve()) if candidate.exists() else str(candidate)
        if key in seen:
            continue
        seen.add(key)
        if candidate.exists():
            return candidate
    return None


def copy_cell_format(src, dst):
    """کپی کردن استایل یک سلول template روی سلول خروجی."""
    if src.has_style:
        dst._style = copy_style(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.alignment:
        dst.alignment = copy_style(src.alignment)
    if src.protection:
        dst.protection = copy_style(src.protection)


def write_excel_from_template(df, template_path, output_path, verbose=True):
    """
    خروجی Excel را بر اساس film-archive-template.xlsx می‌سازد:
    - شیت/هدر/استایل قالب حفظ می‌شود
    - ردیف نمونه حذف می‌شود
    - داده‌ها دقیقاً داخل ستون‌های قالب پر می‌شوند
    """
    if not template_path:
        return None
    if not HAS_OPENPYXL:
        if verbose:
            print("  ⚠️  openpyxl نصب نیست — خروجی با قالب ساخته نشد")
        return None
    if not (HAS_PANDAS and isinstance(df, pd.DataFrame)):
        return None

    try:
        wb = load_workbook(template_path)
    except Exception as exc:
        if verbose:
            print(f"  ⚠️  قالب Excel خوانده نشد ({template_path}): {exc}")
        return None

    ws = wb["Films"] if "Films" in wb.sheetnames else wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h).strip() for h in headers if h is not None and str(h).strip()]
    if not headers:
        if verbose:
            print(f"  ⚠️  قالب Excel ستون header ندارد: {template_path}")
        return None

    template_df = build_dataframe_for_headers(df, headers)
    if template_df is None:
        return None

    # استایل ردیف نمونه (ردیف ۲) را قبل از حذف ذخیره کن.
    style_row = 2 if ws.max_row >= 2 else 1
    row_height = ws.row_dimensions[style_row].height
    source_cells = [ws.cell(row=style_row, column=c) for c in range(1, len(headers) + 1)]

    # حذف داده/نمونه‌های قبلی و پر کردن از صفر.
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    for r_idx, (_, out_row) in enumerate(template_df.iterrows(), start=2):
        if row_height:
            ws.row_dimensions[r_idx].height = row_height
        for c_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = out_row.get(header, "")
            copy_cell_format(source_cells[c_idx - 1], cell)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    if verbose:
        print(f"  🧩 Excel قالب‌دار → {output_path.name} (template: {Path(template_path).name})")
    return template_df


# =============================================================================
# FILE READING
# =============================================================================
def read_input_file(file_path):
    """خواندن فایل ورودی — CSV یا Excel با تشخیص خودکار انکدینگ"""
    ext = Path(file_path).suffix.lower()

    if ext in (".xlsx", ".xls"):
        if not HAS_PANDAS:
            print("  ❌ pandas نصب نیست. لطفاً: pip install pandas openpyxl")
            sys.exit(1)
        df = pd.read_excel(file_path)
        return df

    elif ext == ".csv":
        # تشخیص انکدینگ
        encoding = "utf-8"
        if HAS_CHARDET:
            with open(file_path, "rb") as f:
                raw = f.read(10000)
                detected = chardet.detect(raw)
                if detected and detected.get("encoding"):
                    encoding = detected["encoding"]
                    # اصلاح windows-1252
                    if encoding.lower() in ("windows-1252", "iso-8859-1", "iso-8859-9"):
                        encoding = "windows-1252"

        if HAS_PANDAS:
            try:
                df = pd.read_csv(file_path, encoding=encoding)
            except Exception:
                df = pd.read_csv(file_path, encoding="utf-8", encoding_errors="replace")
            return df
        else:
            # بدون pandas — از csv استفاده کن
            rows = []
            with open(file_path, "r", encoding=encoding, errors="replace") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    rows.append(row)
            return rows
    else:
        print(f"  ❌ فرمت فایل پشتیبانی نمی‌شود: {ext}")
        sys.exit(1)


def find_title_column(df):
    """پیدا کردن ستون عنوان"""
    if HAS_PANDAS and isinstance(df, pd.DataFrame):
        cols = df.columns.tolist()
    else:
        cols = list(df[0].keys()) if df else []

    # اولویت ستون‌ها از دقیق‌ترین تا کلی‌ترین
    candidates = [
        "Title", "title",
        "Movie Name", "movie name", "Movie name",
        "Series Name", "series name", "Series name",
        "Name", "name",
        "عنوان", "نام", "نام فیلم",
    ]
    for candidate in candidates:
        for col in cols:
            if str(col).strip() == candidate:
                return col
    # جستجوی substring
    for col in cols:
        col_lower = str(col).strip().lower()
        if "title" in col_lower or "name" in col_lower or "عنوان" in col_lower:
            return col
    # اولین ستون
    return cols[0] if cols else None


def normalize_columns(df):
    """یکسان‌سازی نام ستون‌ها با پشتیبانی از قالب فارسی/انگلیسی."""
    if HAS_PANDAS and isinstance(df, pd.DataFrame):
        mapping = {}
        used = set()
        for col in df.columns:
            clean = str(col).strip()
            canonical = canonical_column_name(clean)
            # اگر همان ستون استاندارد قبلاً وجود دارد، ستون دوم را duplicate نکن.
            if canonical in used:
                mapping[col] = clean
            else:
                mapping[col] = canonical
                used.add(canonical)
        return df.rename(columns=mapping)
    return df


# =============================================================================
# MAIN PROCESSING
# =============================================================================
def process_file(input_path, file_type="auto", output_dir=None,
                 delay=0.5, translate=True, verbose=True,
                 template_path=None, use_template=True, **kwargs):
    """پردازش اصلی فایل"""

    input_path = Path(input_path)
    if not input_path.exists():
        print(f"❌ فایل یافت نشد: {input_path}")
        return False

    # پوشه خروجی
    if not output_dir:
        output_dir = input_path.parent / "output"
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"  🎬 Cinefilm Enricher v2")
    print(f"  📁 فایل ورودی: {input_path.name}")
    print(f"  📂 پوشه خروجی: {output_dir}")
    print(f"{'='*60}")

    # خواندن فایل
    print(f"\n📖 خواندن فایل...")
    df = read_input_file(input_path)

    if HAS_PANDAS and isinstance(df, pd.DataFrame):
        df = normalize_columns(df)
        
        # تشخیص خودکار ستون مسیر فایل (مثل H:\Movies\xxx)
        title_col = find_title_column(df)
        
        # اگر ستون عنوان حاوی مسیر فایل است، نام فیلم را استخراج کن
        if title_col:
            sample_values = df[title_col].dropna().head(5).tolist()
            if any(is_file_path(v) for v in sample_values):
                print(f"  🛤️  ستون '{title_col}' حاوی مسیر فایل است")
                print(f"  📂 استخراج نام فیلم از مسیر...")
                extracted_col = "Title"
                df[extracted_col] = df[title_col].apply(lambda x: extract_name_from_path(x) if pd.notna(x) else "")
                title_col = extracted_col
                # ستون Folder Paths هم اضافه کن
                if "Folder Paths" in df.columns:
                    pass  # قبلاً وجود دارد
                else:
                    df["Folder Paths"] = ""
        total = len(df)
        print(f"  📊 {total} ردیف | ستون عنوان: '{title_col}'")
    else:
        total = len(df)
        title_col = find_title_column(df)
        print(f"  📊 {total} ردیف | ستون عنوان: '{title_col}'")

    # اگر نام ستون عنوان متفاوت است (مثلاً «عنوان فیلم (Title)»)، ستون استاندارد Title را پر کن.
    if HAS_PANDAS and isinstance(df, pd.DataFrame) and title_col is not None:
        if "Title" not in df.columns:
            df["Title"] = df[title_col]
        else:
            empty_title = df["Title"].isna() | (df["Title"].astype(str).str.strip() == "")
            df.loc[empty_title, "Title"] = df.loc[empty_title, title_col]

    # تشخیص نوع
    if file_type == "auto":
        file_type = detect_file_type(input_path, df)
    print(f"  🎭 نوع: {'🎥 فیلم' if file_type == 'movie' else '📺 سریال'}")

    # بارگذاری کش
    cache_name = "omdb_cache.json" if file_type == "movie" else "tvmaze_cache.json"
    cache = Cache(output_dir / cache_name)
    
    # وارد کردن کش خارجی (در صورت مشخص شدن)
    import_cache_path = kwargs.get("import_cache")
    if import_cache_path:
        imported = cache.import_from(import_cache_path)
        print(f"  📥 {imported} آیتم از کش خارجی وارد شد")

    # ستون‌ها و تابع استخراج
    if file_type == "movie":
        columns = MOVIE_COLUMNS
        fetch_fn = lambda title: fetch_movie_details(title, cache, translate, verbose)
    else:
        columns = SERIES_COLUMNS
        fetch_fn = lambda title: fetch_series_details(title, cache, translate, verbose)

    # قالب Excel نهایی (در صورت وجود)
    resolved_template = resolve_template_path(template_path, input_path, use_template)
    if resolved_template and verbose:
        print(f"  🧩 قالب خروجی: {resolved_template.name}")

    # اطمینان از وجود ستون‌ها
    if HAS_PANDAS and isinstance(df, pd.DataFrame):
        for col in columns:
            if col not in df.columns:
                df[col] = ""
            df[col] = df[col].astype(object)

    # پردازش
    print(f"\n🔄 شروع پردازش...\n")
    success_count = 0
    fail_count = 0
    skip_count = 0

    if HAS_PANDAS and isinstance(df, pd.DataFrame):
        for idx, row in df.iterrows():
            title = row.get(title_col)
            if not title or pd.isna(title) or str(title).strip() == "":
                continue

            title = str(title).strip()
            progress = f"[{idx + 1}/{total}]"

            # چک پردازش قبلی
            dir_val = row.get("Director", "")
            genre_val = row.get("Genre", "")
            if (dir_val and not pd.isna(dir_val) and str(dir_val).strip()
                    and genre_val and not pd.isna(genre_val) and str(genre_val).strip()):
                skip_count += 1
                if verbose:
                    print(f"  {progress} ⏭️  '{title}' — قبلاً پردازش شده")
                continue

            if verbose:
                print(f"  {progress} 🔍 '{title}'")

            details = fetch_fn(title)

            if details:
                success_count += 1
                field_map = {
                    "Original Title": details.get("Original Title", "") or details.get("Title", ""),
                    "Director": details.get("Director", ""),
                    "Cast": details.get("Cast", "") or details.get("Actors", ""),
                    "Year": details.get("Year", ""),
                    "Genre": details.get("Genre", ""),
                    "Rating": details.get("Rating", "") or details.get("imdbRating", ""),
                    "Runtime": clean_runtime(details.get("Runtime", "")),
                    "Country": details.get("Country", ""),
                    "Synopsis": details.get("Synopsis", "") or details.get("Plot", ""),
                }

                poster = details.get("Poster URL", "") or details.get("Poster", "")
                if poster and poster != "N/A":
                    field_map["Poster URL"] = poster

                if file_type == "series":
                    field_map["Seasons Available"] = details.get("Seasons Available", "")

                for col, val in field_map.items():
                    if col in df.columns:
                        df.at[idx, col] = val

                if verbose:
                    rating = field_map.get("Rating", "")
                    year = field_map.get("Year", "")
                    extra = f"  ⭐ {rating}" if rating else ""
                    extra += f"  📅 {year}" if year else ""
                    print(f"    ✅ OK{extra}")
            else:
                fail_count += 1
                if verbose:
                    print(f"    ❌ یافت نشد")

            # ذخیره دوره‌ای کش (هر ۵۰ آیتم)
            if (idx + 1) % 50 == 0:
                cache.save()

            # تأخیر
            if not cache.get(f"{'movie' if file_type == 'movie' else 'series'}:{title}"):
                time.sleep(delay)

        # ذخیره نهایی کش
        cache.save()

    # خروجی‌ها
    print(f"\n💾 ذخیره خروجی‌ها...")

    base = input_path.stem
    # فایل غنی‌شده
    enriched_path = output_dir / f"{base}-enriched.xlsx"
    missing_path = output_dir / f"{base}-missing.csv"

    template_df_out = None
    if HAS_PANDAS and isinstance(df, pd.DataFrame):
        # مرتب‌سازی ستون‌ها
        ordered = [c for c in columns if c in df.columns]
        extra = [c for c in df.columns if c not in columns]
        df_out = df[ordered + extra]

        # اگر قالب film-archive-template.xlsx موجود باشد، Excel نهایی دقیقاً با همان قالب ساخته می‌شود.
        template_df_out = write_excel_from_template(df_out, resolved_template, enriched_path, verbose)
        if template_df_out is None:
            df_out.to_excel(enriched_path, index=False)

        # فایل missing
        missing = df[df["Director"].isna() | (df["Director"] == "") |
                      df["Genre"].isna() | (df["Genre"] == "")]
        if len(missing) > 0:
            missing.to_csv(missing_path, index=False, encoding="utf-8-sig")
            print(f"  ⚠️  {len(missing)} عنوان ناقص → {missing_path.name}")

    print(f"  📊 فایل اصلی → {enriched_path.name}")

    # CSV هم بساز. اگر Excel از قالب ساخته شده، CSV هم با همان ستون‌های قالب نوشته می‌شود.
    csv_path = output_dir / f"{base}-enriched.csv"
    if HAS_PANDAS and isinstance(df, pd.DataFrame):
        csv_df = template_df_out if template_df_out is not None else df_out
        csv_df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        print(f"  📊 CSV → {csv_path.name}")

    # گزارش نهایی
    processed = success_count + fail_count
    print(f"\n{'='*60}")
    print(f"  📈 گزارش نهایی")
    print(f"  {'='*58}")
    print(f"  کل ردیف‌ها:     {total}")
    print(f"  رد شده (قبلی):  {skip_count}")
    print(f"  پردازش شده:     {processed}")
    print(f"  ✅ موفق:         {success_count}")
    print(f"  ❌ ناموفق:       {fail_count}")
    if processed > 0:
        pct = (success_count / processed) * 100
        print(f"  📊 درصد موفقیت: {pct:.1f}%")
    print(f"  📦 کش: {cache.stats()}")
    print(f"{'='*60}\n")

    return True


# =============================================================================
# CLI
# =============================================================================
def main():
    parser = argparse.ArgumentParser(
        description="Cinefilm Archive Enricher v2 — غنی‌سازی خودکار اطلاعات فیلم و سریال",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
مثال‌ها:
  python cinefilm-enricher-v2.py --input movies.csv
  python cinefilm-enricher-v2.py --input series.xlsx --type series
  python cinefilm-enricher-v2.py --input data.csv --type auto --output results/
  python cinefilm-enricher-v2.py --input data.csv --template film-archive-template.xlsx
  python cinefilm-enricher-v2.py --input data.csv --delay 0.5 --no-translate
        """
    )
    parser.add_argument("--input", "-i", required=True, help="فایل ورودی (CSV یا Excel)")
    parser.add_argument("--type", "-t", choices=["auto", "movie", "series"],
                        default="auto", help="نوع فایل (پیش‌فرض: auto)")
    parser.add_argument("--output", "-o", help="پوشه خروجی (پیش‌فرض: input_dir/output)")
    parser.add_argument("--template", help=(
        "مسیر قالب Excel خروجی؛ پیش‌فرض: استفاده خودکار از film-archive-template.xlsx اگر کنار ورودی/برنامه باشد"
    ))
    parser.add_argument("--no-template", action="store_true",
                        help="غیرفعال کردن خروجی قالب‌دار حتی اگر film-archive-template.xlsx موجود باشد")
    parser.add_argument("--delay", "-d", type=float, default=0.5,
                        help="تأخیر بین درخواست‌ها (ثانیه، پیش‌فرض: 0.5)")
    parser.add_argument("--no-translate", action="store_true",
                        help="غیرفعال کردن ترجمه خودکار")
    parser.add_argument("--import-cache", help="وارد کردن کش از فایل JSON خارجی (مثل all_omdb_results.json)")
    parser.add_argument("--export-cache", help="صادر کردن کش به فایل JSON بعد از اجرا")
    parser.add_argument("--quiet", "-q", action="store_true",
                        help="حالت بی‌صدا (فقط خطاها)")

    args = parser.parse_args()

    # بررسی وابستگی‌ها
    missing_deps = []
    if not HAS_PANDAS:
        missing_deps.append("pandas")
    if not HAS_TRANSLATOR and not args.no_translate:
        print("⚠️  deep-translator نصب نیست — ترجمه خودکار غیرفعال")
        print("   نصب: pip install deep-translator\n")

    success = process_file(
        input_path=args.input,
        file_type=args.type,
        output_dir=args.output,
        delay=args.delay,
        translate=not args.no_translate,
        verbose=not args.quiet,
        template_path=args.template,
        use_template=not args.no_template,
        import_cache=args.import_cache,
        export_cache=args.export_cache,
    )
    
    # صادر کردن کش بعد از اجرا (اگر درخواست شده باشد)
    if args.export_cache and success:
        print(f"\n  📤 صادر کردن کش...")
        cache_path = Path(args.output or Path(args.input).parent / "output")
        cache_file = cache_path / ("omdb_cache.json" if args.type != "series" else "tvmaze_cache.json")
        if cache_file.exists():
            import shutil
            shutil.copy2(cache_file, args.export_cache)
            print(f"  ✅ کش صادر شد به: {args.export_cache}")

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
